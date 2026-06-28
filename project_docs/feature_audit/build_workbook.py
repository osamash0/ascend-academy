# -*- coding: utf-8 -*-
"""
Bootstrap the canonical FEATURE_AUDIT.xlsx from audit_data + audit_data2.

Run once to create the workbook. Phases 2-4 update cells in-place via
update_workbook helpers (see update_results.py), so this generator is only
re-run if the Phase-1 story set itself changes.

  python project_docs/feature_audit/build_workbook.py
"""
import os
from collections import Counter, OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import audit_data
import audit_data2

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "FEATURE_AUDIT.xlsx")

ROWS = audit_data.ROWS + audit_data2.ROWS

# Columns: Phase 1 (filled) + Phase 2-4 (empty, filled later)
HEADERS = [
    "ID", "Feature", "Sub-feature", "Role", "User Story",
    "Expected Behavior (from code)", "Key Files", "Code Status (P1)", "Evidence (P1)",
    "Test Result (P2)", "Errors / Notes (P2)", "Severity (P2)",
    "Fix Applied (P3)", "Retest Result (P4)",
]
# widths per column (chars)
WIDTHS = [10, 16, 26, 11, 46, 56, 30, 14, 34, 13, 40, 11, 40, 16]
WRAP_COLS = {5, 6, 7, 9, 11, 13, 14}  # 1-based columns to wrap

STATUS_FILL = {
    "Functioning": "C6EFCE",
    "Partial":     "FFEB9C",
    "Broken":      "FFC7CE",
    "Unclear":     "D9D9D9",
}
STATUS_FONT = {
    "Functioning": "1E6B34",
    "Partial":     "9C6500",
    "Broken":      "9C0006",
    "Unclear":     "555555",
}
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Alignment(vertical="top", wrap_text=True)
TOP_NW = Alignment(vertical="top", wrap_text=False)


def _check_unique_ids():
    ids = [r[0] for r in ROWS]
    dupes = [k for k, v in Counter(ids).items() if v > 1]
    if dupes:
        raise SystemExit(f"Duplicate IDs: {dupes}")
    return ids


def build():
    _check_unique_ids()
    wb = Workbook()

    # ── Sheet 1: Stories ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "User Stories"
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS[c - 1]

    for r in ROWS:
        rid, feat, sub, role, story, expected, files, status, evidence = r
        ws.append([rid, feat, sub, role, story, expected, files, status, evidence,
                   "", "", "", "", ""])
        rownum = ws.max_row
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=rownum, column=c)
            cell.border = BORDER
            cell.alignment = TOP if c in WRAP_COLS else TOP_NW
        # ID bold
        ws.cell(row=rownum, column=1).font = Font(bold=True, size=10)
        # status color
        st = ws.cell(row=rownum, column=8)
        if status in STATUS_FILL:
            st.fill = PatternFill("solid", fgColor=STATUS_FILL[status])
            st.font = Font(bold=True, color=STATUS_FONT[status], size=10)
        st.alignment = Alignment(vertical="top", horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    ws.sheet_view.showGridLines = False

    # ── Sheet 2: Dashboard ───────────────────────────────────────────
    ds = wb.create_sheet("Dashboard")
    ds.sheet_view.showGridLines = False
    ds.column_dimensions["A"].width = 30
    for col in "BCDEF":
        ds.column_dimensions[col].width = 14

    ds["A1"] = "ascend-academy — Feature Audit Dashboard"
    ds["A1"].font = Font(bold=True, size=15, color="1F3864")
    ds["A2"] = f"Total user stories: {len(ROWS)}   |   Branch: abdullah (post v1-restructure merge)"
    ds["A2"].font = Font(italic=True, color="555555")

    # group by feature-area prefix (ID before '-')
    areas = OrderedDict()
    for r in ROWS:
        prefix = r[0].split("-")[0]
        areas.setdefault(prefix, []).append(r[7])  # status

    AREA_NAMES = {
        "AUTH": "Auth / Access / Settings", "PDF": "PDF Upload & Parsing",
        "CRS": "Courses / Enrollment / Library", "AI": "AI Content & Tutor",
        "MAP": "Mind Map / Concept Graph", "WRK": "Worksheets / Practice / Assignments",
        "DSH": "Student Dashboard", "NDG": "Nudges / Planner", "ONB": "Onboarding / Academic",
        "GAM": "Gamification", "RNK": "Ranks", "SOC": "Social / Leaderboard",
        "ANL": "Analytics", "ADM": "Admin", "FBK": "Feedback",
    }

    hdr_row = 4
    ds.cell(row=hdr_row, column=1, value="Feature Area")
    for i, h in enumerate(["Total", "Functioning", "Partial", "Broken", "Unclear"]):
        ds.cell(row=hdr_row, column=2 + i, value=h)
    for c in range(1, 7):
        cell = ds.cell(row=hdr_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    totals = Counter()
    rownum = hdr_row + 1
    for prefix, statuses in areas.items():
        cnt = Counter(statuses)
        totals.update(cnt)
        ds.cell(row=rownum, column=1, value=AREA_NAMES.get(prefix, prefix))
        ds.cell(row=rownum, column=2, value=len(statuses))
        for i, key in enumerate(["Functioning", "Partial", "Broken", "Unclear"]):
            v = cnt.get(key, 0)
            cell = ds.cell(row=rownum, column=3 + i, value=v)
            cell.alignment = Alignment(horizontal="center")
            if v and key in STATUS_FILL:
                cell.fill = PatternFill("solid", fgColor=STATUS_FILL[key])
                cell.font = Font(color=STATUS_FONT[key], bold=(key in ("Broken",)))
        ds.cell(row=rownum, column=1).font = Font(bold=True)
        ds.cell(row=rownum, column=2).alignment = Alignment(horizontal="center")
        rownum += 1

    # totals row
    ds.cell(row=rownum, column=1, value="TOTAL").font = Font(bold=True, size=11)
    ds.cell(row=rownum, column=2, value=len(ROWS)).font = Font(bold=True)
    ds.cell(row=rownum, column=2).alignment = Alignment(horizontal="center")
    for i, key in enumerate(["Functioning", "Partial", "Broken", "Unclear"]):
        cell = ds.cell(row=rownum, column=3 + i, value=totals.get(key, 0))
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
    for c in range(1, 7):
        ds.cell(row=rownum, column=c).fill = PatternFill("solid", fgColor="E8EAF0")

    # legend / phase notes
    note_row = rownum + 3
    ds.cell(row=note_row, column=1, value="Phase tracking").font = Font(bold=True, size=12, color="1F3864")
    notes = [
        "P1 Code Status — derived from reading current code (this pass). DONE.",
        "P2 Test Result — runtime test of each story; PASS / FAIL / BLOCKED. (pending)",
        "P2 Errors/Notes — observed defects with repro. (pending)",
        "P3 Fix Applied — logistical/UX fix landed. (pending)",
        "P4 Retest Result — re-test after fix; PASS / FAIL. (pending)",
    ]
    for i, n in enumerate(notes):
        ds.cell(row=note_row + 1 + i, column=1, value=n)

    wb.save(OUT)
    print(f"Wrote {OUT}  ({len(ROWS)} stories)")
    # quick console summary
    for prefix, statuses in areas.items():
        c = Counter(statuses)
        print(f"  {prefix:5} {len(statuses):3}  F{c.get('Functioning',0)} P{c.get('Partial',0)} B{c.get('Broken',0)} U{c.get('Unclear',0)}")
    print(f"  TOTAL {len(ROWS)}  F{totals.get('Functioning',0)} P{totals.get('Partial',0)} B{totals.get('Broken',0)} U{totals.get('Unclear',0)}")


if __name__ == "__main__":
    build()
