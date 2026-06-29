# -*- coding: utf-8 -*-
"""
Update FEATURE_AUDIT.xlsx in-place with Phase 2-4 results.

Usage: edit the UPDATES dict / TEST_LOG list below and run:
  python project_docs/feature_audit/update_results.py

- UPDATES maps story ID -> dict of any of:
    test (P2), errors (P2), severity (P2), fix (P3), retest (P4)
  Only provided keys are written; existing cells are preserved otherwise.
- TEST_LOG appends rows to a "Test Log" sheet (chronological Phase-2/3/4 notes).
Re-runnable: writing the same values is idempotent.
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
WB = os.path.join(HERE, "FEATURE_AUDIT.xlsx")

# column index (1-based) by phase field
COL = {"test": 10, "errors": 11, "severity": 12, "fix": 13, "retest": 14}

RESULT_FILL = {
    "PASS": "C6EFCE", "FAIL": "FFC7CE", "BLOCKED": "FFEB9C",
    "FIXED": "C6EFCE", "N/A": "D9D9D9",
}
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Alignment(vertical="top", wrap_text=True)

# ─────────────────────────────────────────────────────────────────────────────
# Per-story updates. Append as phases progress.
# ─────────────────────────────────────────────────────────────────────────────
UPDATES = {
    # AI tutor rows — backend was un-importable due to tutor.py:124 orphaned
    # `return False/return True` (IndentationError). Fixed in Phase 2.
    "AI-01": {"test": "BLOCKED→PASS", "errors": "Backend un-importable: tutor.py:124 IndentationError (orphaned dead code from removed refusal helper) broke main + all tutor-importing routers. Import-checks pass after fix.", "severity": "Critical", "fix": "Removed orphaned `return False`/`return True` at tutor.py:123-125."},
    "AI-04": {"test": "PASS", "errors": "_extract_citations compiles+imports clean after tutor.py fix.", "severity": "—"},
    "AI-21": {"test": "PASS (behavior change confirmed)", "errors": "Out-of-scope refusal short-circuit was intentionally removed; the dead `return False/True` remnant caused the syntax error. Stale test test_tutor_grounding::test_out_of_scope_refusal_short_circuits_llm still asserts old behavior.", "severity": "Low", "fix": "Removed dead remnant (same as AI-01). Stale test to be reconciled."},

    # Frontend typecheck — clean across all TS/TSX
    "AUTH-13": {"test": "PASS", "errors": "tsc --noEmit: 0 errors. Verified empirically: backend serves only /api/v1/*, but main.py:150 redirect_legacy_api 307-redirects /api/*→/api/v1/* and fetch follows it preserving method+body, so apiClient (rewrites) AND raw fetch() call sites both resolve. NOT a routing bug.", "severity": "—"},

    # Auth middleware — covered by repaired integration tests (httpx /user path)
    "AUTH-20": {"test": "PASS", "errors": "test_auth_middleware verify_token (httpx GET /user + cache) green after test reconciliation.", "severity": "—"},
    # Backend suite covers these flows green:
    "AI-13": {"test": "PASS", "errors": "test_ai_endpoints TestSummary green (generate_summary via tutor_service).", "severity": "—"},
    "AI-14": {"test": "PASS", "errors": "test_ai_endpoints TestQuiz green.", "severity": "—"},
    "PDF-08": {"test": "PASS", "errors": "test_check_duplicate green; /api/upload/check-duplicate resolves via 307 compat redirect.", "severity": "—"},
    "PDF-14": {"test": "PASS", "errors": "test_v4_quiz_mapping green (answer normalization, bad-answer drop).", "severity": "—"},
    "PDF-01": {"test": "PASS", "errors": "test_unified_orchestrator suite green (happy path, replay, resume, missing-pdf).", "severity": "—"},
    "MAP-02": {"test": "PASS", "errors": "mind_map endpoints import + tests green.", "severity": "—"},
    "CRS-01": {"test": "PASS", "errors": "test_courses_endpoints create/list/get green after fake-supabase timestamp fix.", "severity": "—"},
    "CRS-19": {"test": "PASS", "errors": "test_courses_endpoints student-sees-only-enrolled green.", "severity": "—"},

    # ── Phase 3 fixes ──
    "NDG-12": {"test": "BROKEN (confirmed)", "errors": "OptimalScheduleCard fully built but zero render sites → planner UI unreachable.", "severity": "Medium", "fix": "Mounted <OptimalScheduleCard/> in StudentDashboard below-fold after KnowledgeMapCard (+ import).", "retest": "PASS (tsc clean; mounted). Browser confirm needs auth."},
    "AUTH-05": {"test": "BROKEN (confirmed)", "errors": "Password-reset email redirects to /reset-password which had NO route → user landed on protected 404; reset impossible in-app.", "severity": "High", "fix": "Added src/pages/ResetPassword.tsx (handles Supabase recovery session via onAuthStateChange+getSession, updateUser({password}), signs out, →/auth; invalid/expired-link + done states). Registered BARE route /reset-password in App.tsx (not PublicRoute, which would bounce the recovery session). Added PublicRoutes.RESET_PASSWORD.", "retest": "PASS (tsc 0 errors; route resolves). Full recovery-link flow needs live Supabase email."},
    "SOC-13": {"test": "Partial (confirmed)", "errors": "Leaderboard period toggle labeled 'Daily'/'Monthly' but mapped to week(last-7d)/all(all-time).", "severity": "Low", "fix": "Relabeled buttons 'This Week' / 'All Time' to match the underlying period semantics (logic unchanged).", "retest": "PASS (tsc clean)."},
    "PDF-02": {"test": "Partial (confirmed)", "errors": ".env=5 (live) but config default + .env.example said 2 → fresh deploy w/o env var silently ran old v2 pipeline.", "severity": "Medium", "fix": "Set config.py parser_version default '2'→'5' and .env.example PARSER_VERSION 2→5 to match live.", "retest": "PASS (backend 638 green with default=5)."},
    "AI-20": {"test": "Partial (confirmed)", "errors": "ai_content._AiModelLiteral and slides_ai._AiModel diverged (slides_ai had unsupported 'deepseek'; ai_content had 'llama3' + lacked gemini-1.5/gpt-4o-mini) → a model accepted by one AI router rejected (422) by the other.", "severity": "Low", "fix": "Aligned both literals to orchestrator._USER_MODEL_TO_PROVIDER's 13 supported keys; dropped unsupported deepseek/llama3 (they silently fell back anyway; frontend only sends 4).", "retest": "PASS (backend 638 green)."},
    "GAM-12": {"test": "Partial (confirmed)", "errors": "Achievements.tsx locked 'Potential Milestones' built from a hard-coded possibleAchievements array, not the DB badge_definitions catalog → drifts from real catalog (Ascent.tsx does it right).", "severity": "Low", "fix": "Refactored Achievements.tsx to fetchBadgeCatalog() and derive locked = catalog minus earned (by name) minus is_secret. Removed the hard-coded array.", "retest": "PASS (tsc clean; vitest 219)."},
    "GAM-15": {"test": "Partial (confirmed)", "errors": "Achievements.tsx rendered raw English badge_name (no i18n), unlike Ascent.tsx (badgeLabel).", "severity": "Low", "fix": "Localized earned + locked badge names via badgeLabel(t, def), resolving each earned row's catalog key by name.", "retest": "PASS (tsc clean; vitest 219)."},

    # ── Phase 3 batch 2 ──
    "AUTH-15": {"test": "Partial (confirmed)", "errors": "Two divergent guard impls: src/lib/routeGuards.tsx (PublicRoute omits admin redirect) was dead — App.tsx uses its own inline, more-complete guards. Hazard: importing the wrong one.", "severity": "Low", "fix": "Deleted unused src/lib/routeGuards.tsx + its test (only covered the dead version). App.tsx inline guards remain the single source.", "retest": "PASS (tsc 0; vitest 209 = 219-10 deleted-dead-code tests)."},
    "AUTH-34": {"test": "Partial (confirmed)", "errors": "SettingsWrapper had a role branch whose two arms returned identical JSX (dead branching).", "severity": "Low", "fix": "Collapsed to a single ConsoleLayout<Settings/> return.", "retest": "PASS (tsc 0)."},
    "AUTH-38": {"test": "Partial (confirmed)", "errors": "ProtectedNotFound professor/non-professor branches returned identical JSX.", "severity": "Low", "fix": "Collapsed to authed→shell / unauthed→bare NotFound.", "retest": "PASS (tsc 0)."},
    "ANL-12": {"test": "Partial (confirmed)", "errors": "Student Questions Feed subtitle said '(Grouped by similarity)' but backend returns flat chronological — false claim.", "severity": "Low", "fix": "Relabeled to '(most recent first)' to match actual behavior.", "retest": "PASS (tsc 0)."},
    "PDF-29": {"test": "Partial (confirmed)", "errors": "Parser selector labeled Auto as 'v4 Recommended' though live pipeline is v5.", "severity": "Low", "fix": "Relabeled to 'Auto (Recommended)' (version-agnostic).", "retest": "PASS (tsc 0)."},
    "PDF-23": {"test": "Partial (confirmed)", "errors": "Fast-upload hardcoded model='gemini/gemini-2.0-flash' in 3 functions, ignoring the configurable-LLM design.", "severity": "Low", "fix": "Added settings.fast_upload_model (FAST_UPLOAD_MODEL, litellm-format, default unchanged = gemini/gemini-2.0-flash) and referenced it in all 3 acompletion calls.", "retest": "PASS (backend 638; import OK)."},
    "CRS-27": {"test": "Partial (confirmed)", "errors": "Library course sort hardcoded /database/i.test(title) to force 'Database Systems' first — demo bias.", "severity": "Low", "fix": "Removed the bias; sort is now has-lectures → last-opened → progress. Updated stale comment.", "retest": "PASS (tsc 0; vitest 209)."},
    "CRS-23": {"test": "Partial (confirmed)", "errors": "Course details sheet showed FABRICATED ratings (4.8 / 124) + generic 'what you'll learn' when real data absent — misleading.", "severity": "Low", "fix": "Dropped fake fallbacks (averageRating→undefined, ratingCount→0, outcomes→[]); CourseDetailsSheet only renders the rating block when ratingCount>0 && averageRating!=null.", "retest": "PASS (tsc 0; vitest 209)."},
    "AI-08": {"test": "Partial (confirmed)", "errors": "LectureChat sent Accept: text/event-stream and had a full SSE branch, but backend /ai/chat only returns JSON → SSE path was dead.", "severity": "Low", "fix": "Removed the dead SSE branch + reader loop; kept the JSON path; Accept→application/json. finally{} still resets streamingRef.", "retest": "PASS (tsc 0; vitest LectureChat green)."},

    # ── Deferred / documented (not silently 'fixed') ──
    "PDF-04": {"test": "BROKEN (confirmed)", "errors": "'Skip AI'/on_demand toggle is a no-op under live v5 (unified branch never reads parsing_mode; always full-AI synth).", "severity": "Medium", "fix": "DEFERRED — needs product decision: (a) honor on_demand in the unified pipeline (skip synthesis) = real pipeline work, or (b) hide the toggle when v5 is live (frontend can't currently see PARSER_VERSION). Not silently changed."},
    "PDF-05": {"test": "Partial (confirmed)", "errors": "use_blueprint only threaded into the legacy parser; ignored by v4/v5.", "severity": "Low", "fix": "DEFERRED — dev/Pipeline-Lab-only option; harmless no-op. Left as-is (documenting rather than removing a tester control)."},
    "PDF-27": {"test": "Partial (confirmed)", "errors": "Pipeline Test Lab reads slide._meta + question.answer (v4 shape); v5 emits flat slides + correctAnswer index → route/layout/token stats empty, deck correct-option highlight wrong.", "severity": "Low", "fix": "DEFERRED — professor-only internal dev inspector, not a student/prof production surface. Tracked for a v5-aware rework."},

    # ── Phase 3 batch 3 (this pass): two remaining genuine logistical errors ──
    "PDF-10": {"test": "Partial (confirmed)", "errors": "Cache-invalidation guard omitted v5: at PARSER_VERSION=5 a v4-shaped parse cache could be replayed instead of running the unified pipeline.", "severity": "Medium", "fix": "upload.py: map auto→'unified' at v5 and add 'unified' to the parser-mismatch invalidation set so a non-unified cache is dropped before replay.", "retest": "PASS (check_duplicate suite green; backend 703 non-db)."},
    "AUTH-29": {"test": "Partial (confirmed)", "errors": "Account deletion removed client-reachable rows but NOT the Supabase auth user → auth identity (and non-client-reachable data) persisted. GDPR erasure gap.", "severity": "High", "fix": "New POST /api/auth/delete-account (verify_token, 3/min) calls supabase_admin.auth.admin.delete_user(uid) → cascades via auth.users FKs; invalidates token cache first. Settings.handleDelete calls it authoritatively, falling back to client row-deletes if unavailable. +3 backend tests (mocked admin API).", "retest": "PASS (backend tests green; tsc 0). NOTE: live admin-delete + FK cascade need verification against a real Supabase project."},

    # ── Dispositioned non-defects (config / feature-gap / by-design) — NOT code errors ──
    "PDF-15": {"fix": "BY DESIGN — documented", "retest": "v5 generates a deck-level quiz only; per-slide quizzes are authored in the editor. Intentional."},
    "PDF-16": {"fix": "KNOWN GAP — documented", "retest": "Client auto-suggested quizzes not persisted on the v5 server-authoritative save path. Needs save-path rework; deferred (Tier 2+)."},
    "CRS-18": {"fix": "FEATURE GAP — documented", "retest": "Unenroll endpoint exists; no UI calls it. Missing feature, not an error."},
    "AI-02":  {"fix": "CONFIG — documented", "retest": "Zero-vector embeddings only when GEMINI_API_KEY unset (tutor degrades to current-slide grounding). Ops config, not a code defect."},
    "WRK-14": {"fix": "BY DESIGN — documented", "retest": "free_form self-assessed (excluded from score); short_answer exact-match. Pinned in test_practice_sheet_grading; intentional 'for now'."},
    "NDG-08": {"fix": "CONFIG — documented", "retest": "Nudge scheduler gated behind ENABLE_NUDGE_SCHEDULER (ops choice)."},
    "NDG-13": {"fix": "FEATURE GAP — documented", "retest": "Backend honors study_minutes_per_day; no UI to set it yet."},
    "ONB-09": {"fix": "BY DESIGN — documented", "retest": "Course recs surfaced in the library catalog sheet; dashboard placement is optional polish."},
    "SOC-05": {"fix": "FEATURE GAP — documented", "retest": "remove_friend RPC + mutation wired; no remove-friend button in audited pages."},
    "SOC-16": {"fix": "DEAD CODE — documented", "retest": "bootstrap_demo_friends is a demo-seeding util with no UI caller by design."},
    "ANL-16": {"fix": "INTENTIONAL — documented", "retest": "Predictive/Spatial panels feature-flagged off in code on purpose."},
    "ANL-22": {"fix": "MINOR — documented", "retest": "Real metrics with canned headlines; /insights intentionally redirects to the Intelligence Center."},
    "ADM-02": {"fix": "FEATURE GAP — documented", "retest": "Admin user view is read-only; role management (endpoint+UI) not built. Future feature."},
    "ADM-05": {"fix": "CONFIG — documented", "retest": "Sentry diagnostics show mock data unless SENTRY_* env set; the UI banner says so."},
    "ADM-08": {"fix": "COSMETIC — documented", "retest": "app_version/pool sizes are hardcoded display strings; non-functional."},
    "FBK-03": {"fix": "FEATURE GAP — documented", "retest": "Feedback is write-only; no admin review surface yet. Future feature."},
}

# ─────────────────────────────────────────────────────────────────────────────
# Test Log — chronological Phase-2+ entries (area, what, result, detail)
# ─────────────────────────────────────────────────────────────────────────────
TEST_LOG = [
    ("2026-06-24", "P2", "BUILD: Frontend typecheck", "PASS", "npx tsc --noEmit -p tsconfig.json → 0 errors. Frontend compiles clean after the v1-restructure merge."),
    ("2026-06-24", "P2", "BUILD: Backend import-check", "FAIL→FIXED", "Importing backend.main + all 17 v1 routers initially failed: 6 modules hit IndentationError at services/ai/tutor.py:124 (orphaned return False/return True from a removed refusal helper) → entire backend un-importable. Removed the dead lines → all 18 modules import OK."),
    ("2026-06-24", "P2", "TEST: Backend suite (pytest -m 'not db')", "617/677 PASS", "After tutor.py fix + conftest patch repair: 617 passed, 12 failed, 9 errors. Failures/errors are predominantly test-infra staleness from the v1 restructure (monkeypatch targets moved API→service modules; stubs missing new kwargs) plus 4 diagnostics tests doing real network calls. No NEW product regressions identified yet."),
    ("2026-06-24", "P2", "FIX: conftest patch_supabase", "FIXED", "ai_content (v1) no longer exposes create_client (uses supabase_admin; AI create_client moved to tutor_service). Repointed conftest to patch database.create_client + ai_content.supabase_admin. Cleared the global 'ai_api has no attribute create_client' setup error."),
    ("2026-06-24", "P2", "TRIAGE: stale integration patches", "OPEN", "test_ai_endpoints patches ai_content.generate_summary (moved→tutor_service); test_check_duplicate patches upload.validate_upload (moved→upload_service); test_file_parse_service stub missing vision_model kwarg; test_auth_middleware mocks supabase_admin.auth.get_user (verify_token now uses httpx GET /user). To reconcile in Phase 3 as restructure test-debt."),
    ("2026-06-24", "P3", "FIX: backend test-debt reconciled", "638 PASS", "Reconciled all 21 stale backend tests (repointed monkeypatch targets to service modules, updated stubs for vision_model kwarg + httpx verify_token, updated removed-refusal test, added fake_supabase created_at/updated_at, mocked diagnostics network). Backend suite: 638 passed / 0 failed / 0 errors. No product bugs masked."),
    ("2026-06-24", "P2", "TEST: Frontend suite (vitest)", "202/219→219/219", "Initial: 17 failed across 4 files (apiClient, analyticsService, professorOverview, LectureChat) — all from the /api→/api/v1 apiClient rewrite (stale MSW handler URLs) + removed tutor refusal text. Reconciled via dual /api+/api/v1 MSW handler registration + updated LectureChat mock. Frontend suite: 219 passed / 0 failed."),
    ("2026-06-24", "P2", "VERIFY: /api vs /api/v1 routing", "NOT A BUG", "Backend mounts only /api/v1/*. Suspected ~15 raw fetch(`/api/...`) call sites (NudgeBanner, useTTS, usePDFUpload, useLectureSubmit, FastUpload, conceptsService, worksheetsService) would 404. Ran uvicorn + curl: /api/* returns 307 → /api/v1/* via main.py:150 redirect_legacy_api; fetch follows 307 preserving method+body. Both conventions resolve. Minor: extra round-trip; watch cross-origin Authorization on redirect (same-origin via nginx is fine)."),
    ("2026-06-24", "P2", "ENV: disk space", "RESOLVED", "Disk hit ~98% mid-run (one ENOSPC during vitest). Cleared vite/vitest/coverage caches (~4.7 GiB). Now ~72% used."),
    ("2026-06-24", "P3", "FIX: undeclared dep python-json-logger", "FIXED", "New untracked backend/core/logging_config.py (imported by main.py:17) needs pythonjsonlogger, but it was NOT in requirements.txt → fresh `pip install -r requirements.txt` deploy would fail to start the backend. It also went missing locally mid-session (disk cleanup). Reinstalled + added 'python-json-logger>=3.0' to backend/requirements.txt."),
    ("2026-06-24", "P3", "FIX: DomainError handler crash", "FIXED", "REAL BUG (surfaced once logging_config loaded): main.py:179 logged with extra={'msg': ...} — 'msg' is a reserved LogRecord attribute → logging raised KeyError, crashing the DomainError exception handler itself (every domain error would 500 in the handler). Renamed key to 'error_message'. test_api_v1_structure::test_domain_error_exception_handler now green."),
    ("2026-06-24", "P3", "FIX batch (logistical/UX)", "5 FIXED", "SOC-13 leaderboard labels (This Week/All Time); PDF-02 parser_version default 2→5 (config + .env.example); AUTH-05 reset-password page+route; AI-20 unified ai_model allow-lists; NDG-12 planner card mounted (prior). All confirmed against the relevant Phase-1 finding before fixing."),
    ("2026-06-24", "P4", "RETEST: full suites + typecheck", "ALL GREEN", "After the Phase-3 batch: backend pytest 638 passed / 0 fail; frontend vitest 219 passed / 0 fail; tsc --noEmit 0 errors. (Runtime/browser confirmation of AUTH-05 recovery-link + NDG-12 render still needs a live authenticated stack.)"),
    ("2026-06-24", "P2", "SCOPE: non-defect Partials", "DOCUMENTED", "Some Partial/Unclear items are config- or feature-scope, NOT code defects to 'fix': AI-02 (needs GEMINI_API_KEY), ADM-05 (Sentry mock unless SENTRY_* env), ADM-02 (admin role-management not built), ADM-08 (hardcoded app_version), FBK-03 (no feedback-review surface), NDG-08/13 (scheduler env flag / no budget UI). Recorded as known limitations rather than bug-fixed."),
    ("2026-06-24", "P3", "FIX batch 2 (logistical/UX)", "8 FIXED", "AUTH-15 (deleted dead routeGuards+test); AUTH-34/38 (collapsed identical-branch JSX); ANL-12 (truthful feed label); PDF-29 (version-agnostic parser label); PDF-23 (configurable FAST_UPLOAD_MODEL); CRS-27 (removed Database-first sort bias); CRS-23 (no fabricated ratings — render only with real data); AI-08 (removed dead SSE branch from LectureChat)."),
    ("2026-06-24", "P3", "DEFER: 3 items documented", "DEFERRED", "PDF-04 (Skip-AI no-op under v5 — needs product decision: honor on_demand vs hide toggle); PDF-05 (dead use_blueprint, dev-lab only); PDF-27 (Pipeline Test Lab v4-meta drift, professor-only dev inspector). Not silently changed; tracked in the workbook."),
    ("2026-06-24", "P4", "RETEST: full suites + typecheck (final batch)", "ALL GREEN", "After batch 2: backend pytest 638 passed; frontend vitest 209 passed (was 219 − 10 deleted dead-code tests); tsc --noEmit 0 errors. All landed Phase-3 fixes retested green."),
    ("2026-06-29", "P3", "FIX batch 3 (remaining genuine errors)", "2 FIXED", "PDF-10 (v5 added to cache-invalidation guard — stale v4 cache no longer replayed under v5); AUTH-29 (new POST /api/auth/delete-account using service-role admin.delete_user → GDPR cascade erasure; Settings wired best-effort with the old client-row-delete as fallback)."),
    ("2026-06-29", "P3", "DISPOSITION: 16 non-defects", "DOCUMENTED", "Every remaining Broken/Partial item is now dispositioned in the workbook: config (AI-02, NDG-08, ADM-05), feature-gap (CRS-18, NDG-13, SOC-05, ADM-02, FBK-03), by-design (PDF-15, WRK-14, ONB-09, ANL-16), dead-code/cosmetic/minor (SOC-16, ADM-08, ANL-22), known-gap-deferred (PDF-16). None are code defects to fix."),
    ("2026-06-29", "P3", "TEST SUITE: Tier-1 expansion", "54 NEW", "Built per the approved test plan: file_validation (13), practice_sheet grading (10), idempotency (6), upload_service routing (4), slides_ai endpoints (10), practice_sheets access-control (5), + practice_sheets RLS db tests (6). Consolidated the duplicate usePDFUpload test file (8 tests, 2→1). Coverage: backend 67%→69%; file_validation/idempotency 100%; slides_ai 16→68%."),
    ("2026-06-29", "P4", "RETEST: FULL post-fix (all phases)", "GREEN + 2 pre-existing", "Backend non-db 703 passed / 0 fail; backend db 45 passed / 0 fail; tsc 0 errors; frontend 386 passed / 2 failed. The 2 frontend failures are in Onboarding.test.tsx — caused by an uncommitted teammate WIP (native <select>→Radix <Select>) out of sync with the committed test; NOT from any audit change. Flagged for the owner."),
]


def main():
    wb = load_workbook(WB)
    ws = wb["User Stories"]

    # index rows by ID
    id_row = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            id_row[str(v).strip()] = r

    written = 0
    missing = []
    for sid, fields in UPDATES.items():
        r = id_row.get(sid)
        if not r:
            missing.append(sid)
            continue
        for key, val in fields.items():
            c = COL[key]
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = TOP
            cell.border = BORDER
            # color the test/retest result cell by leading token
            if key in ("test", "retest", "fix"):
                token = str(val).replace("→", " ").split()[0].upper() if val else ""
                fill = RESULT_FILL.get(token) or (RESULT_FILL.get("FIXED") if key == "fix" and val else None)
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)
            written += 1

    # Test Log sheet
    if "Test Log" in wb.sheetnames:
        ls = wb["Test Log"]
    else:
        ls = wb.create_sheet("Test Log")
        ls.append(["Date", "Phase", "Item", "Result", "Detail"])
        for c in range(1, 6):
            hc = ls.cell(row=1, column=c)
            hc.fill = PatternFill("solid", fgColor="1F3864")
            hc.font = Font(bold=True, color="FFFFFF")
        for col, w in zip("ABCDE", [12, 7, 34, 16, 95]):
            ls.column_dimensions[col].width = w
        ls.freeze_panes = "A2"
        ls.sheet_view.showGridLines = False

    # de-dupe: only append log rows not already present (by Item+Detail)
    existing = set()
    for r in range(2, ls.max_row + 1):
        existing.add((ls.cell(row=r, column=3).value, ls.cell(row=r, column=5).value))
    for row in TEST_LOG:
        if (row[2], row[4]) in existing:
            continue
        ls.append(list(row))
        rn = ls.max_row
        for c in range(1, 6):
            ls.cell(row=rn, column=c).alignment = TOP
        res = ls.cell(row=rn, column=4)
        token = str(row[3]).replace("→", " ").split()[-1].upper() if row[3] else ""
        if token in RESULT_FILL:
            res.fill = PatternFill("solid", fgColor=RESULT_FILL[token])

    wb.save(WB)
    print(f"Updated {written} story cells across {len(UPDATES)} stories; Test Log now {ls.max_row-1} entries.")
    if missing:
        print("WARNING missing IDs:", missing)


if __name__ == "__main__":
    main()
